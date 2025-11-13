# generation.py
import pandas as pd
import streamlit as st
from datetime import datetime
from fpdf import FPDF
import base64
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

# Roman numeral conversion
def int_to_roman(num):
    """Convert integer to Roman numeral"""
    val = [10, 9, 5, 4, 1]
    syms = ['X', 'IX', 'V', 'IV', 'I']
    roman_num = ''
    i = 0
    while num > 0:
        for _ in range(num // val[i]):
            roman_num += syms[i]
            num -= val[i]
        i += 1
    return roman_num

def format_subject_for_excel(row):
    """
    Format subject for Excel display:
    [CM Group] Subject Name - (Module Code) [Difficulty Tag]
    """
    subject = str(row.get('Subject', ''))
    
    # CM Group prefix
    cm_group = str(row.get('CMGroup', '')).strip()
    if cm_group and cm_group not in ['', 'nan', 'None']:
        cm_group_prefix = f"[{cm_group}] "
    else:
        cm_group_prefix = ""
    
    # Difficulty suffix (only for COMP category)
    difficulty_suffix = ""
    if row.get('Category') == 'COMP' and pd.notna(row.get('Difficulty')):
        difficulty = row.get('Difficulty')
        if difficulty == 0 or difficulty == '0':
            difficulty_suffix = " (Easy)"
        elif difficulty == 1 or difficulty == '1':
            difficulty_suffix = " (Difficult)"
    
    return cm_group_prefix + subject + difficulty_suffix

def format_elective_for_excel(row):
    """
    Format elective subject for Excel display:
    [CM Group] Subject Name - (Module Code) [OE Type] [Difficulty Tag]
    """
    subject = str(row.get('Subject', ''))
    oe_type = str(row.get('OE', ''))
    
    # CM Group prefix
    cm_group = str(row.get('CMGroup', '')).strip()
    if cm_group and cm_group not in ['', 'nan', 'None']:
        cm_group_prefix = f"[{cm_group}] "
    else:
        cm_group_prefix = ""
    
    # OE type
    oe_suffix = f" [{oe_type}]" if oe_type and oe_type not in ['', 'nan', 'None'] else ""
    
    # Difficulty suffix
    difficulty_suffix = ""
    if row.get('Category') == 'COMP' and pd.notna(row.get('Difficulty')):
        difficulty = row.get('Difficulty')
        if difficulty == 0 or difficulty == '0':
            difficulty_suffix = " (Easy)"
        elif difficulty == 1 or difficulty == '1':
            difficulty_suffix = " (Difficult)"
    
    return cm_group_prefix + subject + oe_suffix + difficulty_suffix


def save_to_excel(sem_dict, college_name):
    """Generate main timetable Excel with EXACT formatting."""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sem, df in sem_dict.items():
            # Get main branch and semester info
            main_branch = df['MainBranch'].iloc[0] if not df.empty else 'B TECH'
            semester_roman = int_to_roman(sem)
            
            # Separate non-electives and electives
            df_non_elec = df[~(df['OE'].notna() & (df['OE'].str.strip() != ''))].copy()
            df_elec = df[df['OE'].notna() & (df['OE'].str.strip() != '')].copy()
            
            # Process non-electives
            if not df_non_elec.empty:
                sheet_name = f"{main_branch}_Sem_{semester_roman}"
                
                # Format subjects
                df_non_elec['SubjectDisplay'] = df_non_elec.apply(format_subject_for_excel, axis=1)
                
                # Convert dates
                df_non_elec['Exam Date'] = pd.to_datetime(df_non_elec['Exam Date'], format="%d-%m-%Y", errors='coerce')
                df_non_elec = df_non_elec.dropna(subset=['Exam Date'])
                df_non_elec = df_non_elec.sort_values('Exam Date')
                
                # Create pivot table
                pivot_df = df_non_elec.groupby(['Exam Date', 'SubBranch'])['SubjectDisplay'].apply(
                    lambda x: ', '.join(sorted(x))
                ).reset_index()
                
                pivot_table = pivot_df.pivot(index='Exam Date', columns='SubBranch', values='SubjectDisplay')
                pivot_table = pivot_table.fillna('---')
                
                # Format dates back to DD-MM-YYYY
                pivot_table.index = pivot_table.index.strftime('%d-%m-%Y')
                pivot_table.index.name = 'Exam Date'
                
                # Write to Excel
                pivot_table.to_excel(writer, sheet_name=sheet_name)
                
                # Style the sheet
                worksheet = writer.sheets[sheet_name]
                
                # Header row (row 1)
                for col_num, col_name in enumerate(pivot_table.columns, 2):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.value = col_name
                    cell.font = Font(name="Arial", size=11, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="951C1C", end_color="951C1C", fill_type="solid")
                    cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                
                # Exam Date header
                cell = worksheet.cell(row=1, column=1)
                cell.value = "Exam Date"
                cell.font = Font(name="Arial", size=11, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="951C1C", end_color="951C1C", fill_type="solid")
                cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                
                # Column widths
                worksheet.column_dimensions['A'].width = 15
                for col in range(2, len(pivot_table.columns) + 2):
                    worksheet.column_dimensions[get_column_letter(col)].width = 40
                
                # Data rows styling
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    for cell in row:
                        cell.font = Font(name="Arial", size=10)
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.border = thin_border
                        
                        # Alternating row colors
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Process electives
            if not df_elec.empty:
                sheet_name = f"{main_branch}_Sem_{semester_roman}_Electives"
                
                # Format subjects
                df_elec['SubjectDisplay'] = df_elec.apply(format_elective_for_excel, axis=1)
                
                # Convert dates
                df_elec['Exam Date'] = pd.to_datetime(df_elec['Exam Date'], format="%d-%m-%Y", errors='coerce')
                df_elec = df_elec.dropna(subset=['Exam Date'])
                df_elec = df_elec.sort_values('Exam Date')
                
                # Group by date and OE type
                elec_grouped = df_elec.groupby(['Exam Date', 'OE'])['SubjectDisplay'].apply(
                    lambda x: ', '.join(sorted(x))
                ).reset_index()
                
                # Format dates
                elec_grouped['Exam Date'] = elec_grouped['Exam Date'].dt.strftime('%d-%m-%Y')
                elec_grouped.columns = ['Exam Date', 'OE Type', 'Subjects']
                
                # Write to Excel
                elec_grouped.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Style the sheet
                worksheet = writer.sheets[sheet_name]
                
                # Header row styling
                for col_num in range(1, 4):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="951C1C", end_color="951C1C", fill_type="solid")
                
                # Column widths
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 10
                worksheet.column_dimensions['C'].width = 80
                
                # Data rows styling
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    for cell in row:
                        cell.font = Font(name="Arial", size=10)
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        cell.border = thin_border
                        
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    output.seek(0)
    return output


def save_verification_excel(sem_dict, college_name):
    """Generate verification Excel (same format as main)."""
    return save_to_excel(sem_dict, college_name)


def generate_pdf_timetable(sem_dict, college_name):
    """Generate PDF timetable with EXACT layout."""
    
    class PDF(FPDF):
        def __init__(self, orientation='L', unit='mm', format='A4'):
            # Custom page size: 210mm x 500mm for landscape with extended width
            super().__init__(orientation=orientation, unit=unit, format=(210, 500))
            self.alias_nb_pages()
            self.header_content = None
            self.branches = []
            self.time_slot = None
            self.logo_path = None
            
        def header(self):
            if self.header_content:
                # Logo (if exists)
                logo_x = (self.w - 40) / 2
                if self.logo_path and os.path.exists(self.logo_path):
                    self.image(self.logo_path, x=logo_x, y=10, w=40)
                    y_start = 30
                else:
                    y_start = 10
                
                # Red banner with school name
                self.set_fill_color(149, 33, 28)
                self.set_text_color(255, 255, 255)
                self.set_font("Arial", 'B', 16)
                banner_height = 14
                self.rect(10, y_start, self.w - 20, banner_height, 'F')
                self.set_xy(10, y_start)
                self.cell(self.w - 20, banner_height, college_name.upper(), 0, 1, 'C')
                
                y_start += banner_height + 2
                
                # Program and semester
                self.set_font("Arial", 'B', 15)
                self.set_text_color(0, 0, 0)
                self.set_xy(10, y_start)
                program_text = f"{self.header_content['main_branch']} - Semester {self.header_content['semester_roman']}"
                self.cell(self.w - 20, 8, program_text, 0, 1, 'C')
                
                y_start += 8
                
                # Time slot
                if self.time_slot:
                    self.set_font("Arial", 'B', 14)
                    self.set_xy(10, y_start)
                    self.cell(self.w - 20, 6, f"Exam Time: {self.time_slot}", 0, 1, 'C')
                    y_start += 6
                
                # Check time note
                self.set_font("Arial", 'I', 10)
                self.set_xy(10, y_start)
                self.cell(self.w - 20, 6, "(Check the subject exam time)", 0, 1, 'C')
                y_start += 6
                
                # Branches
                self.set_font("Arial", '', 12)
                self.set_xy(10, y_start)
                branches_text = f"Branches: {', '.join(self.branches)}"
                self.cell(self.w - 20, 6, branches_text, 0, 1, 'C')
                
                self.ln(5)
        
        def footer(self):
            self.set_y(-20)
            self.set_font("Arial", 'B', 14)
            self.cell(0, 5, "Controller of Examinations", 0, 1, 'L')
            self.line(10, self.get_y(), 70, self.get_y())
            
            # Page number
            self.set_y(-15)
            self.set_font("Arial", '', 14)
            page_text = f"Page {self.page_no()} of {{nb}}"
            self.cell(0, 5, page_text, 0, 0, 'R')
    
    def get_preferred_slot(semester):
        """Get preferred time slot based on semester"""
        if semester % 2 != 0:  # Odd semester
            odd_sem_position = (semester + 1) // 2
            return "10:00 AM - 1:00 PM" if odd_sem_position % 2 == 1 else "2:00 PM - 5:00 PM"
        else:  # Even semester
            even_sem_position = semester // 2
            return "10:00 AM - 1:00 PM" if even_sem_position % 2 == 1 else "2:00 PM - 5:00 PM"
    
    def format_date_long(date_str):
        """Convert DD-MM-YYYY to 'Monday, 1 April, 2025'"""
        try:
            date_obj = datetime.strptime(date_str, "%d-%m-%Y")
            return date_obj.strftime("%A, %d %B, %Y")
        except:
            return date_str
    
    def wrap_text(pdf, text, max_width):
        """Wrap text to fit within max_width"""
        words = str(text).split()
        lines = []
        current_line = ""
        
        for word in words:
            test_line = word if not current_line else current_line + " " + word
            if pdf.get_string_width(test_line) <= max_width:
                current_line = test_line
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
        
        if current_line:
            lines.append(current_line)
        
        return lines if lines else [str(text)]
    
    pdf = PDF(orientation='L')
    pdf.set_auto_page_break(auto=True, margin=25)
    
    for sem, df in sem_dict.items():
        if df.empty:
            continue
        
        main_branch = df['MainBranch'].iloc[0]
        semester_roman = int_to_roman(sem)
        time_slot = get_preferred_slot(sem)
        
        # Separate non-electives and electives
        df_non_elec = df[~(df['OE'].notna() & (df['OE'].str.strip() != ''))].copy()
        df_elec = df[df['OE'].notna() & (df['OE'].str.strip() != '')].copy()
        
        # Process non-electives
        if not df_non_elec.empty:
            pdf.header_content = {
                'main_branch': main_branch,
                'semester_roman': semester_roman
            }
            pdf.branches = sorted(df_non_elec['SubBranch'].unique().tolist())
            pdf.time_slot = time_slot
            
            pdf.add_page()
            
            # Format subjects
            df_non_elec['SubjectDisplay'] = df_non_elec.apply(format_subject_for_excel, axis=1)
            df_non_elec['Exam Date'] = pd.to_datetime(df_non_elec['Exam Date'], format="%d-%m-%Y", errors='coerce')
            df_non_elec = df_non_elec.dropna(subset=['Exam Date'])
            df_non_elec = df_non_elec.sort_values('Exam Date')
            
            # Create pivot
            pivot_df = df_non_elec.groupby(['Exam Date', 'SubBranch'])['SubjectDisplay'].apply(
                lambda x: ', '.join(sorted(x))
            ).reset_index()
            
            pivot_table = pivot_df.pivot(index='Exam Date', columns='SubBranch', values='SubjectDisplay')
            pivot_table = pivot_table.fillna('---')
            
            # Calculate column widths
            date_col_width = 60
            num_branches = len(pivot_table.columns)
            remaining_width = pdf.w - 20 - date_col_width
            branch_col_width = remaining_width / num_branches
            
            col_widths = [date_col_width] + [branch_col_width] * num_branches
            
            # Table header
            pdf.set_fill_color(149, 33, 28)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", 'B', 10)
            
            pdf.cell(col_widths[0], 10, "Exam Date", 1, 0, 'C', True)
            for idx, branch in enumerate(pivot_table.columns, 1):
                pdf.cell(col_widths[idx], 10, str(branch), 1, 0, 'C', True)
            pdf.ln()
            
            # Table data
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", '', 10)
            
            row_num = 0
            for date_idx, row in pivot_table.iterrows():
                date_str = date_idx.strftime('%d-%m-%Y')
                date_long = format_date_long(date_str)
                
                # Wrap date text
                date_lines = wrap_text(pdf, date_long, col_widths[0] - 4)
                
                # Wrap subject texts
                subject_lines = []
                max_lines = len(date_lines)
                
                for branch in pivot_table.columns:
                    subject_text = row[branch]
                    lines = wrap_text(pdf, subject_text, col_widths[1] - 4)
                    subject_lines.append(lines)
                    max_lines = max(max_lines, len(lines))
                
                # Calculate row height
                line_height = 5
                row_height = line_height * max_lines
                
                # Alternating row color
                if row_num % 2 == 1:
                    pdf.set_fill_color(240, 240, 240)
                else:
                    pdf.set_fill_color(255, 255, 255)
                
                # Draw cells
                x_start = pdf.get_x()
                y_start = pdf.get_y()
                
                # Date cell
                for line_idx, line in enumerate(date_lines):
                    pdf.set_xy(x_start, y_start + line_idx * line_height)
                    pdf.cell(col_widths[0], line_height, line, 0, 0, 'C', True)
                
                pdf.rect(x_start, y_start, col_widths[0], row_height)
                
                # Subject cells
                for col_idx, lines in enumerate(subject_lines, 1):
                    x_pos = x_start + sum(col_widths[:col_idx])
                    for line_idx, line in enumerate(lines):
                        pdf.set_xy(x_pos, y_start + line_idx * line_height)
                        pdf.cell(col_widths[col_idx], line_height, line, 0, 0, 'C', True)
                    
                    pdf.rect(x_pos, y_start, col_widths[col_idx], row_height)
                
                pdf.set_xy(x_start, y_start + row_height)
                pdf.ln()
                row_num += 1
        
        # Process electives
        if not df_elec.empty:
            pdf.header_content = {
                'main_branch': main_branch + " (Electives)",
                'semester_roman': semester_roman
            }
            pdf.branches = ["Open Electives"]
            pdf.time_slot = time_slot
            
            pdf.add_page()
            
            # Format subjects
            df_elec['SubjectDisplay'] = df_elec.apply(format_elective_for_excel, axis=1)
            df_elec['Exam Date'] = pd.to_datetime(df_elec['Exam Date'], format="%d-%m-%Y", errors='coerce')
            df_elec = df_elec.dropna(subset=['Exam Date'])
            df_elec = df_elec.sort_values('Exam Date')
            
            # Group by date and OE
            elec_grouped = df_elec.groupby(['Exam Date', 'OE'])['SubjectDisplay'].apply(
                lambda x: ', '.join(sorted(x))
            ).reset_index()
            
            # Column widths
            date_col_width = 60
            oe_col_width = 30
            subject_col_width = pdf.w - 20 - date_col_width - oe_col_width
            
            col_widths = [date_col_width, oe_col_width, subject_col_width]
            
            # Table header
            pdf.set_fill_color(149, 33, 28)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", 'B', 10)
            
            pdf.cell(col_widths[0], 10, "Exam Date", 1, 0, 'C', True)
            pdf.cell(col_widths[1], 10, "OE Type", 1, 0, 'C', True)
            pdf.cell(col_widths[2], 10, "Subjects", 1, 0, 'C', True)
            pdf.ln()
            
            # Table data
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", '', 10)
            
            row_num = 0
            for _, row in elec_grouped.iterrows():
                date_str = row['Exam Date'].strftime('%d-%m-%Y')
                date_long = format_date_long(date_str)
                oe_type = row['OE']
                subjects = row['SubjectDisplay']
                
                # Wrap texts
                date_lines = wrap_text(pdf, date_long, col_widths[0] - 4)
                oe_lines = [oe_type]
                subject_lines = wrap_text(pdf, subjects, col_widths[2] - 4)
                
                max_lines = max(len(date_lines), len(oe_lines), len(subject_lines))
                line_height = 5
                row_height = line_height * max_lines
                
                # Alternating row color
                if row_num % 2 == 1:
                    pdf.set_fill_color(240, 240, 240)
                else:
                    pdf.set_fill_color(255, 255, 255)
                
                # Draw cells
                x_start = pdf.get_x()
                y_start = pdf.get_y()
                
                # Date cell
                for line_idx, line in enumerate(date_lines):
                    pdf.set_xy(x_start, y_start + line_idx * line_height)
                    pdf.cell(col_widths[0], line_height, line, 0, 0, 'C', True)
                pdf.rect(x_start, y_start, col_widths[0], row_height)
                
                # OE Type cell
                pdf.set_xy(x_start + col_widths[0], y_start)
                pdf.cell(col_widths[1], row_height, oe_type, 0, 0, 'C', True)
                pdf.rect(x_start + col_widths[0], y_start, col_widths[1], row_height)
                
                # Subjects cell
                for line_idx, line in enumerate(subject_lines):
                    pdf.set_xy(x_start + col_widths[0] + col_widths[1], y_start + line_idx * line_height)
                    pdf.cell(col_widths[2], line_height, line, 0, 0, 'C', True)
                pdf.rect(x_start + col_widths[0] + col_widths[1], y_start, col_widths[2], row_height)
                
                pdf.set_xy(x_start, y_start + row_height)
                pdf.ln()
                row_num += 1
    
    pdf_output = BytesIO()
    pdf_output.write(pdf.output(dest='S').encode('latin-1'))
    pdf_output.seek(0)
    return pdf_output


def create_download_link(file_data, filename, file_type):
    """Create styled download button."""
    b64 = base64.b64encode(file_data.read()).decode()
    file_data.seek(0)
    return f'''
        <a href="data:application/octet-stream;base64,{b64}" download="{filename}">
            <button style="
                background-color: #4CAF50; color: white; padding: 10px 20px;
                text-align: center; text-decoration: none; display: inline-block;
                font-size: 14px; margin: 4px 2px; cursor: pointer; border: none; border-radius: 4px;
            ">
                Download {file_type}
            </button>
        </a>
    '''
